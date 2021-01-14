# 监控或记录PLC的标签值，并且可以通过excel导出
# 2021/12/05 wei he
# *********************************************
# -*- coding: utf-8 -*-
import sys

from PyQt5 import uic
from PyQt5.QtCore import QTimer
from PyQt5.QtWidgets import QApplication
from openpyxl import Workbook
from pylogix import PLC

# 实例化 plc
AB = PLC()
# 定义一些全局变量
old_tag = 0
i = 0
old_string = ""
j = 0
ID_list = []
String_ID_list = []
dint_list = []
string_list = []


def bytes_to_int(byte):
    """byte 转换成 INT"""
    result = 0
    for b in byte:
        result = result * 256 + int(b)
    return result


def PLC_connect(IP):
    """建立与PLC的连接"""
    AB.IPAddress = str(IP)


def show_string(string_tag):
    """显示字符串内容"""
    tag = AB.Read(string_tag)
    # 字符串标签长度，转换为整型
    length = bytes_to_int(tag.Value[0:1])
    # 字符串内容提取
    # data = bytes(tag.Value[4:length + 4]).decode('unicode_escape')
    data = bytes(tag.Value[4:length + 4]).decode('utf-8')
    # print(data)
    # print('String length is : ' + str(length))
    return data, length


def show_tag(tag):
    """显示标签数值，实数或者整数"""
    local_tag = AB.Read(tag)
    # print(local_tag.Value)
    return local_tag.Value


def information_show(control, msg):
    """软件信息显示"""
    control.append(msg)
    cursor = control.textCursor()
    control.moveCursor(cursor.End)
    QApplication.processEvents()


def run_monitor_tags(monitor_tag, int_select, string_select, control):
    """运行PLC标签监控程序"""
    global i, j, old_tag, old_string, dint_list, string_list, ID_list, String_ID_list
    if int_select:
        # 读取PLC签
        check_tag = show_tag(monitor_tag)
        if old_tag == check_tag:
            pass
        else:
            old_tag = check_tag
            i += 1
            ID_list.append(i)
            dint_list.append(check_tag)
            msg1 = "Int/Float  " + str(i) + " : " + str(check_tag)
            information_show(control, msg1)
    elif string_select:
        # 读取plc字符串标签
        check_string = show_string(monitor_tag)
        if old_string == check_string[0]:
            pass
        else:
            old_string = check_string[0]
            # 显示出字符串内容
            j += 1
            String_ID_list.append(j)
            string_list.append(check_string[0])
            msg2 = "String " + str(j) + " : " + check_string[0] + ', len:' + str(check_string[1])
            information_show(control, msg2)


def export_excel(file_path, monitor_tag, id_list, data_list):
    """导出到excel文件中"""
    # 创建一个工作表，写只读会导致崩溃
    # wb = Workbook(write_only=True)
    wb = Workbook()
    # ws操作sheet页
    ws = wb.active
    ws1 = wb.create_sheet('PLC Tag Data', 0)
    ws1['A1'] = 'ID NUM'
    ws1['B1'] = 'DATA'
    ws1['C1'] = 'Tag Name'
    ws1['C2'] = monitor_tag
    for i in range(2, len(id_list) + 1):
        ws1.cell(row=i, column=1).value = id_list[i - 2]
        ws1.cell(row=i, column=2).value = data_list[i - 2]
    wb.save(file_path)


# 主程序
# ******************************************
class Main_window:

    def __init__(self):
        # 从文件中加载UI定义
        self.ui = uic.loadUi("PLC_tag_read.ui")
        # 默认开始类型选择整型
        self.ui.Dint_float_select.setChecked(True)
        self.ui.string_select.setChecked(False)
        self.ui.monitor_start.setEnabled(False)
        self.ui.monitor_stop.setEnabled(False)
        self.ui.save_excel.setEnabled(False)
        self.ui.excel_path.setText('D:/plc_tag_data.xlsx')
        self.ui.IP_input.setText('10.126.41.98')
        self.ui.Tag_input.setText('hw_python_dint')
        #  初始化 变量值
        global i, j, old_string, old_tag
        old_tag = 0
        i = 0
        old_string = ""
        j = 0
        # 初始化自定义变量，获取UI的内容
        self.input_tag = ''
        self.dint_mode_select = self.ui.Dint_float_select.isChecked()
        self.string_mode_select = self.ui.string_select.isChecked()
        # self.save_button_clicked = self.ui.save_excel.isChecked()
        self.save_button_clicked = False
        # 监控按钮按下，开始读取数据
        self.ui.monitor_start.clicked.connect(self.auto_fresh_tag)
        # 按下关闭按钮停止监控
        self.ui.shut_down.clicked.connect(self.close_soft)
        # 按下停止/导出按钮，停止监控
        self.ui.monitor_stop.clicked.connect(self.stop_auto_tag)
        self.ui.save_excel.clicked.connect(self.stop_auto_tag)
        # 按下导出键，导出excel文件
        self.ui.save_excel.clicked.connect(self.export_data)
        # 定时刷新1.检查IP地址和标签有无输入
        self.time1 = QTimer()
        self.time1.start(100)
        self.time1.timeout.connect(self.button_enable)
        # 运行按钮监控程序
        self.button_enable()
        # 定义定时刷新2，用于实时监控PLC 标签值内容变化
        self.time2 = QTimer()

    def button_enable(self):
        """定时刷新界面检查IP和tag 有无内容,有内容了才显示监控按钮"""
        if self.ui.IP_input.text() != "" and self.ui.Tag_input.text() != "":
            self.ui.monitor_start.setEnabled(True)
            self.time1.stop()

    def auto_fresh_tag(self):
        """自动定时刷新标签监控函数"""
        msg1 = 'Tag read running!'
        information_show(self.ui.data_display, msg1)
        self.ui.Tag_input.setReadOnly(True)
        self.ui.IP_input.setReadOnly(True)
        self.ui.Dint_float_select.setEnabled(False)
        self.ui.string_select.setEnabled(False)
        self.time2.start(50)
        self.time2.timeout.connect(self.view_tag)

    def stop_auto_tag(self):
        """停止标签监控程序"""
        self.time2.stop()
        self.ui.monitor_start.setEnabled(True)
        self.ui.monitor_stop.setEnabled(False)
        self.ui.Tag_input.setReadOnly(False)
        self.ui.IP_input.setReadOnly(False)
        self.ui.Dint_float_select.setEnabled(True)
        self.ui.string_select.setEnabled(True)
        # print(ID_list)
        # print(dint_list)
        # print(string_list)

    def view_tag(self):
        """监控PLC标签"""
        # 输入IP地址, 建立连接
        ip_address = self.ui.IP_input.text()
        PLC_connect(ip_address)
        self.input_tag = self.ui.Tag_input.text()
        self.dint_mode_select = self.ui.Dint_float_select.isChecked()
        self.string_mode_select = self.ui.string_select.isChecked()
        self.save_button_clicked = self.ui.save_excel.isChecked()
        # 调用PLC标签监控函数
        run_monitor_tags(self.input_tag, self.ui.Dint_float_select.isChecked(),
                         self.ui.string_select.isChecked(),
                         self.ui.data_display)
        # 设置按钮可用/不可用
        self.ui.monitor_start.setEnabled(False)
        self.ui.monitor_stop.setEnabled(True)
        self.ui.save_excel.setEnabled(True)

    def export_data(self):
        """导出数据到excel"""
        try:
            file_name = self.ui.excel_path.text()
            tag = self.ui.Tag_input.text()
            if self.ui.Dint_float_select.isChecked():
                export_excel(file_name, tag, ID_list, dint_list)
            elif self.ui.string_select.isChecked():
                export_excel(file_name, tag, String_ID_list, string_list)
            msg = 'Excel export successfully !'
            information_show(self.ui.data_display, msg)
        except:
            err_infor = 'Excel export failed ! Please check if the excel is opened! If so,please close it firstly!'
            information_show(self.ui.data_display, err_infor)

    def close_soft(self):
        """退出系统"""
        AB.Close()
        self.ui.close()
        quit()


# 运行窗口
if __name__ == "__main__":
    # 固定的，PyQt5程序都需要QApplication对象。sys.argv是命令行参数列表，确保程序可以双击运行
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    # 初始化
    myWin = Main_window()
    # 将窗口控件显示在屏幕上
    myWin.ui.show()
    # 程序运行，sys.exit死循环监控屏幕。
    sys.exit(app.exec_())
